# routes/import_export.py — импорт из Excel
import re
from datetime import date, datetime, timedelta

import openpyxl
from flask import Blueprint, flash, redirect, render_template, request, url_for
from openpyxl.utils.exceptions import InvalidFileException

from models import Athlete, ChangeLog, db

bp = Blueprint("import_export", __name__)

HEADER_MARKER = "№ п/п"
FIELD_NAMES = [
    "discipline", "category", "age_category", "rank",
    "organization", "territory", "trainer",
    "result_regional", "result_national", "result_international",
    "national_team",
]


def _clean(value):
    return value.strip() if isinstance(value, str) else value


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=value)).date()
    if isinstance(value, str) and value.strip():
        try:
            return datetime.strptime(value.strip(), "%d.%m.%Y").date()
        except ValueError:
            return None
    return None


def _split_full_name(full_name):
    parts = full_name.split(None, 2)
    last_name = parts[0] if len(parts) > 0 else ""
    first_name = parts[1] if len(parts) > 1 else ""
    middle_name = parts[2] if len(parts) > 2 else None
    return last_name, first_name, middle_name


def _find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(isinstance(c, str) and c.strip() == HEADER_MARKER for c in row):
            return i
    return None


def _row_to_athlete_fields(row):
    """15-колоночная строка данных -> словарь полей Athlete (без ФИО)."""
    last_name, first_name, middle_name = _split_full_name(_clean(row[1]))
    values = [
        _clean(row[4]), _clean(row[5]), _clean(row[6]), _clean(row[7]),
        _clean(row[8]), _clean(row[9]), _clean(row[10]),
        _clean(row[11]), _clean(row[12]), _clean(row[13]), _clean(row[14]),
    ]
    fields = dict(zip(FIELD_NAMES, values))
    fields.update(
        last_name=last_name,
        first_name=first_name,
        middle_name=middle_name,
        birth_date=_to_date(row[2]),
        gender=_clean(row[3]),
    )
    return fields


def _iter_data_rows(ws, start_row):
    """Строки данных начиная с start_row (пустые/служебные строки — включая пустые
    разделители и вторую строку заголовка — пропускаются, а не обрывают импорт)."""
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        name = row[1] if len(row) > 1 else None
        if isinstance(name, str) and name.strip():
            yield row


def _load_sheet(file_storage):
    try:
        wb = openpyxl.load_workbook(file_storage, data_only=True)
    except (InvalidFileException, OSError, KeyError):
        raise ValueError("Файл повреждён или не является Excel-файлом (.xlsx)")
    sheet_name = "Список" if "Список" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError('Не найдена строка заголовка ("№ п/п") — проверьте формат файла')
    return ws, header_row


def _find_existing(fields):
    return Athlete.query.filter_by(
        last_name=fields["last_name"],
        first_name=fields["first_name"],
        middle_name=fields["middle_name"],
        birth_date=fields["birth_date"],
    ).first()


@bp.route("/import")
def import_index():
    return render_template("import/index.html", today=date.today().isoformat())


@bp.route("/import/changes")
def import_changes_index():
    return render_template("import/index.html", today=date.today().isoformat())


@bp.route("/import/upload", methods=["POST"])
def import_upload():
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Выберите файл для загрузки", "error")
        return redirect(url_for("import_export.import_index"))

    try:
        ws, header_row = _load_sheet(file)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("import_export.import_index"))

    added, updated, skipped = 0, 0, 0
    for row in _iter_data_rows(ws, header_row + 1):
        fields = _row_to_athlete_fields(row)
        if fields["birth_date"] is None:
            skipped += 1
            continue
        athlete = _find_existing(fields)
        if athlete:
            for key, value in fields.items():
                setattr(athlete, key, value)
            updated += 1
        else:
            db.session.add(Athlete(**fields))
            added += 1

    db.session.commit()
    summary = f"Импорт завершён: добавлено {added}, обновлено {updated}"
    if skipped:
        summary += f", пропущено (не удалось распознать дату рождения) {skipped}"
    flash(summary, "success")
    return redirect(url_for("import_export.import_index"))


@bp.route("/import/changes/upload", methods=["POST"])
def import_changes_upload():
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Выберите файл для загрузки", "error")
        return redirect(url_for("import_export.import_changes_index"))

    try:
        ws, header_row = _load_sheet(file)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("import_export.import_changes_index"))

    doc_date_raw = request.form.get("doc_date", "")
    doc_date = datetime.strptime(doc_date_raw, "%Y-%m-%d").date() if doc_date_raw else date.today()

    doc_number = request.form.get("document_number", "").strip()
    if not doc_number:
        match = re.search(r"(\d+)(?=\.\w+$)", file.filename)
        doc_number = match.group(1) if match else ""

    excluded, included, not_found, skipped = 0, 0, 0, 0
    section = None

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        marker = row[0].strip().lower() if isinstance(row[0], str) else ""
        if marker in ("исключить", "включить"):
            section = "exclude" if marker == "исключить" else "include"
            continue

        name = row[1] if len(row) > 1 else None
        if not (isinstance(name, str) and name.strip()) or section is None:
            continue

        fields = _row_to_athlete_fields(row)
        if fields["birth_date"] is None:
            skipped += 1
            continue
        athlete = _find_existing(fields)

        if section == "exclude":
            if athlete:
                athlete.is_active = False
                db.session.add(ChangeLog(
                    athlete_id=athlete.id, change_type="исключён",
                    change_date=doc_date, document_number=doc_number,
                ))
                excluded += 1
            else:
                not_found += 1
        else:  # include
            if not athlete:
                athlete = Athlete(**fields)
                db.session.add(athlete)
                db.session.flush()
            athlete.is_active = True
            db.session.add(ChangeLog(
                athlete_id=athlete.id, change_type="включён",
                change_date=doc_date, document_number=doc_number,
            ))
            included += 1

    db.session.commit()
    summary = f"Импорт изменений завершён: исключено {excluded}, включено {included}"
    if not_found:
        summary += f", не найдено в базе {not_found}"
    if skipped:
        summary += f", пропущено (не удалось распознать дату рождения) {skipped}"
    flash(summary, "success")
    return redirect(url_for("import_export.import_changes_index"))
