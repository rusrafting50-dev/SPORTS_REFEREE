# excel_generator.py — генерация итоговых документов по образцу
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Times New Roman"
THIN = Side(style="thin")
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
BORDER_BOTTOM = Border(bottom=THIN)
BORDER_TOP = Border(top=THIN)

CATEGORY_PATTERN = re.compile(r"\s*(\(\s*\d+\s*-\s*\d+\s*категория\s*\))", re.IGNORECASE)
DISCIPLINE_SEPARATOR = re.compile(r",\s*")


def _break_discipline(value):
    """Каждая дисциплина через запятую — на новую строку; внутри — перенос перед '(N-N категория)'."""
    if not value:
        return value
    parts = DISCIPLINE_SEPARATOR.split(value)
    parts = [CATEGORY_PATTERN.sub(r"\n\1", part) for part in parts]
    return ",\n".join(parts)

COLUMN_WIDTHS = {
    "A": 5.86, "B": 27.43, "C": 12.71, "D": 12.71, "E": 26.14, "F": 19.0, "G": 15.0,
    "H": 13.0, "I": 30.29, "J": 24.0, "K": 22.86, "L": 9.29, "M": 10.86,
    "N": 12.43, "O": 23.0,
}

# В шапке вид спорта выводится в дательном падеже ("по спортивному туризму на ... год")
SPORT_NAME_DATIVE = {
    "спортивный туризм": "спортивному туризму",
}


def _sport_name_dative(sport_name):
    if not sport_name:
        return sport_name
    return SPORT_NAME_DATIVE.get(sport_name.strip().lower(), sport_name)

COLUMN_HEADERS = [
    "№ п/п",
    "Фамилия, имя, \nотчество (при наличии)",
    "Дата рождения",
    "Пол",
    "Спортивная дисциплина или группа дисциплин",
    "Категория (спортсмен, тренер, специалист)",
    "Возрастная категория в соответствии с ЕВСК",
    "Спортивное звание, разряд",
    "Физкультурно-спортивная организация",
    "Территориальная принадлежность",
    "Личный тренер",
]

BOLD_12 = Font(name=FONT_NAME, size=12, bold=True)
PLAIN_12 = Font(name=FONT_NAME, size=12)
CAPTION_10 = Font(name=FONT_NAME, size=10)
DATA_FONT = Font(name=FONT_NAME, size=10)
DATA_FONT_12 = Font(name=FONT_NAME, size=12)

CENTER = Alignment(horizontal="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_TOP_WRAP = Alignment(horizontal="center", vertical="top", wrap_text=True)
CENTER_BOTTOM_WRAP = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right")
LEFT = Alignment(horizontal="left")
VERTICAL_TEXT = Alignment(horizontal="center", vertical="bottom", wrap_text=True, textRotation=90)


def _set_column_widths(ws):
    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width


def _merge_value(ws, cell_range, value, font, alignment):
    start, _, end = cell_range.partition(":")
    if end and end != start:
        ws.merge_cells(cell_range)
    cell = ws[start]
    cell.value = value
    cell.font = font
    cell.alignment = alignment
    return cell


def _set_border_range(ws, cell_range, border):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


def _write_counters(ws, label_range, label, count_cell, count, bold=False, count_alignment=CENTER_WRAP):
    font = BOLD_12 if bold else PLAIN_12
    _merge_value(ws, label_range, label, font, RIGHT)
    cell = ws[count_cell]
    cell.value = count
    cell.font = font
    cell.alignment = count_alignment


def _count_by_category(athletes):
    trainers = sum(1 for a in athletes if a.category in ("Тренер", "Главный тренер"))
    specialists = sum(1 for a in athletes if a.category == "Специалист")
    sportsmen = sum(1 for a in athletes if a.category == "Спортсмен")
    return trainers, specialists, sportsmen


def _write_athletes_header(ws, header_row, header_font=BOLD_12):
    """Заголовок таблицы (2 строки) + строка нумерации столбцов. Возвращает первую строку данных."""
    for col in range(1, 12):
        letter = get_column_letter(col)
        _merge_value(
            ws, f"{letter}{header_row}:{letter}{header_row + 1}",
            COLUMN_HEADERS[col - 1], header_font, CENTER_WRAP,
        )

    _merge_value(
        ws, f"L{header_row}:N{header_row}",
        "Высший результат сезона на официальных спортивных соревнованиях",
        header_font, CENTER_WRAP,
    )
    for letter, text in (("L", "Московские областные"), ("M", "Всероссийские"), ("N", "Международные")):
        cell = ws[f"{letter}{header_row + 1}"]
        cell.value = text
        cell.font = header_font
        cell.alignment = VERTICAL_TEXT

    _merge_value(
        ws, f"O{header_row}:O{header_row + 1}",
        "Состав спортивной сборной команды Российской Федерации",
        header_font, CENTER_WRAP,
    )

    numbers_row = header_row + 2
    for col in range(1, 16):
        cell = ws.cell(row=numbers_row, column=col, value=col)
        cell.font = PLAIN_12
        cell.alignment = CENTER_WRAP

    for row in range(header_row, numbers_row + 1):
        for col in range(1, 16):
            ws.cell(row=row, column=col).border = BORDER_ALL

    ws.row_dimensions[header_row].height = 49.5
    ws.row_dimensions[header_row + 1].height = 93.75
    ws.row_dimensions[numbers_row].height = 15

    return numbers_row + 1


def _write_athlete_row(ws, row, index, athlete, font=DATA_FONT):
    values = [
        index, athlete.full_name, athlete.birth_date, athlete.gender,
        _break_discipline(athlete.discipline), athlete.category, athlete.age_category, athlete.rank,
        athlete.organization, athlete.territory, athlete.trainer,
        athlete.result_regional, athlete.result_national, athlete.result_international,
        athlete.national_team,
    ]
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.border = BORDER_ALL
        cell.alignment = LEFT_WRAP if col == 2 else CENTER_WRAP
        if col == 3 and value:
            cell.number_format = "dd.mm.yyyy"
    # Высота не задаётся явно: при открытии Excel сам подбирает высоту строки
    # под перенесённый текст (wrap_text), так что длинные значения не обрезаются.


def generate_report(athletes, settings, doc_date):
    """Формирует итоговый список кандидатов в сборную команду МО."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Список"
    _set_column_widths(ws)

    year = doc_date.year if doc_date else (settings.year if settings else None)
    trainers, specialists, sportsmen = _count_by_category(athletes)
    total = len(athletes)

    org_name = settings.organization_name if settings else ""
    chairman = settings.chairman_name if settings else ""
    head_coach = settings.head_coach_name if settings else ""
    sport_name = settings.sport_name if settings else ""

    ws.row_dimensions[1].height = 15.75
    ws.row_dimensions[3].height = 35.25
    ws.row_dimensions[4].height = 12.0
    ws.row_dimensions[5].height = 38.25
    ws.row_dimensions[6].height = 33.0
    ws.row_dimensions[10].height = 15.75

    _merge_value(ws, "B2:G2", "СФОРМИРОВАН", BOLD_12, CENTER)

    _merge_value(ws, "B3:G3", org_name, BOLD_12, CENTER_BOTTOM_WRAP)
    _set_border_range(ws, "B3:G3", BORDER_BOTTOM)
    _merge_value(ws, "B4:G4", "наименование организации", CAPTION_10, CENTER_WRAP)
    _set_border_range(ws, "B4:G4", BORDER_TOP)

    ws["B5"] = "Председатель"
    ws["B5"].font = PLAIN_12
    ws["B5"].alignment = CENTER
    _merge_value(ws, "C5:D5", None, PLAIN_12, CENTER_WRAP)
    _set_border_range(ws, "C5:D5", BORDER_BOTTOM)
    ws.merge_cells("F5:G5")
    ws["E5"] = chairman
    ws["E5"].font = PLAIN_12
    _write_counters(ws, "L5:M5", "Всего человек:", "N5", total, bold=True, count_alignment=CENTER_BOTTOM_WRAP)

    ws["B6"] = "(должность)"
    ws["B6"].font = CAPTION_10
    ws["B6"].alignment = CENTER_TOP_WRAP
    ws["B6"].border = BORDER_TOP
    _merge_value(ws, "C6:D6", "(подпись)", CAPTION_10, CENTER_TOP_WRAP)
    _set_border_range(ws, "C6:D6", BORDER_TOP)
    ws["E6"] = "(фамилия, инициалы)"
    ws["E6"].font = CAPTION_10
    ws["E6"].alignment = CENTER_TOP_WRAP
    ws["E6"].border = BORDER_TOP
    _merge_value(
        ws, "F6:J6",
        "СПИСОК\nкандидатов в спортивную сборную команду Московской области",
        BOLD_12, CENTER_TOP_WRAP,
    )
    _write_counters(ws, "L6:M6", "из них:", "N6", None)

    ws["B7"] = "Главный тренер"
    ws["B7"].font = PLAIN_12
    ws["B7"].alignment = CENTER
    _merge_value(ws, "C7:D7", None, PLAIN_12, CENTER_WRAP)
    _set_border_range(ws, "C7:D7", BORDER_BOTTOM)
    ws["E7"] = head_coach
    ws["E7"].font = PLAIN_12
    ws["F7"] = "по"
    ws["F7"].font = BOLD_12
    ws["F7"].alignment = CENTER
    _merge_value(ws, "G7:I7", _sport_name_dative(sport_name), BOLD_12, CENTER_WRAP)
    _set_border_range(ws, "G7:I7", BORDER_BOTTOM)
    ws["J7"] = f"на {year} год" if year else ""
    ws["J7"].font = BOLD_12
    ws["J7"].alignment = CENTER
    _write_counters(ws, "L7:M7", "тренеры:", "N7", trainers)

    _merge_value(ws, "C8:D8", "(подпись)", CAPTION_10, CENTER_TOP_WRAP)
    _set_border_range(ws, "C8:D8", BORDER_TOP)
    ws["E8"] = "(фамилия, инициалы)"
    ws["E8"].font = CAPTION_10
    ws["E8"].alignment = CENTER_TOP_WRAP
    ws["E8"].border = BORDER_TOP
    _merge_value(ws, "G8:I8", "(наименование вида спорта)", CAPTION_10, CENTER_TOP_WRAP)
    _set_border_range(ws, "G8:I8", BORDER_TOP)
    _write_counters(ws, "L8:M8", "специалисты:", "N8", specialists)

    _write_counters(ws, "L9:M9", "спортсмены:", "N9", sportsmen)

    first_data_row = _write_athletes_header(ws, 11, header_font=PLAIN_12)
    for i, athlete in enumerate(athletes, start=1):
        _write_athlete_row(ws, first_data_row + i - 1, i, athlete, font=DATA_FONT_12)

    return wb


def generate_changes_report(exclude_list, include_list, settings, doc_date, doc_number):
    """Формирует документ «Изменения в список» с разделами Исключить / Включить."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Список"
    _set_column_widths(ws)

    year = doc_date.year if doc_date else (settings.year if settings else None)
    sport_name = settings.sport_name if settings else ""

    ex_trainers, ex_specialists, ex_sportsmen = _count_by_category(exclude_list)
    in_trainers, in_specialists, in_sportsmen = _count_by_category(include_list)

    ws.row_dimensions[1].height = 20.25
    ws.row_dimensions[2].height = 12.0
    ws.row_dimensions[3].height = 23.25
    ws.row_dimensions[4].height = 33.0

    _merge_value(
        ws, "F4:J4",
        "ИЗМЕНЕНИЯ В СПИСОК\nкандидатов в спортивную сборную команду Московской области",
        BOLD_12, CENTER_TOP_WRAP,
    )
    _write_counters(ws, "K3", "Всего исключено:", "L3", len(exclude_list), bold=True, count_alignment=LEFT)
    _write_counters(ws, "N3", "Всего включено:", "O3", len(include_list), bold=True, count_alignment=LEFT)

    ws["F5"] = "по"
    ws["F5"].font = BOLD_12
    ws["F5"].alignment = CENTER
    _merge_value(ws, "G5:I5", _sport_name_dative(sport_name), BOLD_12, CENTER_WRAP)
    _set_border_range(ws, "G5:I5", BORDER_BOTTOM)
    ws["J5"] = f"на {year} год" if year else ""
    ws["J5"].font = BOLD_12
    ws["J5"].alignment = CENTER

    _write_counters(ws, "K4", "из них:", "L4", None, count_alignment=LEFT)
    _write_counters(ws, "N4", "из них:", "O4", None, count_alignment=LEFT)
    _write_counters(ws, "K5", "тренеры:", "L5", ex_trainers, count_alignment=LEFT)
    _write_counters(ws, "N5", "тренеры:", "O5", in_trainers, count_alignment=LEFT)
    _write_counters(ws, "K6", "специалисты:", "L6", ex_specialists, count_alignment=LEFT)
    _write_counters(ws, "N6", "специалисты:", "O6", in_specialists, count_alignment=LEFT)
    _write_counters(ws, "K7", "спортсмены:", "L7", ex_sportsmen, count_alignment=LEFT)
    _write_counters(ws, "N7", "спортсмены:", "O7", in_sportsmen, count_alignment=LEFT)

    _merge_value(ws, "G6:I6", "(наименование вида спорта)", CAPTION_10, CENTER_TOP_WRAP)
    _set_border_range(ws, "G6:I6", BORDER_TOP)

    row = _write_athletes_header(ws, 9, header_font=PLAIN_12)

    _merge_value(ws, f"A{row}:O{row}", "Исключить", BOLD_12, CENTER)
    for cell in ws[row]:
        cell.border = BORDER_ALL
    row += 1
    for i, athlete in enumerate(exclude_list, start=1):
        _write_athlete_row(ws, row, i, athlete, font=DATA_FONT_12)
        row += 1

    include_marker_row = row
    _merge_value(ws, f"A{include_marker_row}:O{include_marker_row}", "Включить", BOLD_12, CENTER)
    for cell in ws[include_marker_row]:
        cell.border = BORDER_ALL
    row += 1
    for i, athlete in enumerate(include_list, start=1):
        _write_athlete_row(ws, row, i, athlete, font=DATA_FONT_12)
        row += 1

    return wb
